import os
import re
import sys
import argparse
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# -------- Configuración --------
START_ROW = 26                     # Fila donde empiezan los encabezados de la tabla del acta
LOCATION_MODE = "raw"              # "raw" = deja completo lo capturado; "first_token" = toma solo la primera palabra (p.ej. "AMAZONAS")
ACTA_MODE = "prefix"               # "prefix" = "ACTA No. 243"; "number_only" = "243"

# -------- Utilidades --------
def norm_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def norm_serial(x):
    s = norm_str(x)
    return re.sub(r"\s+", "", s).upper()

def improved_find_acta_meta_xlsx(path):
    """
    Extrae:
      - date_str (YYYY-MM-DD) desde celdas superiores (o patrón dd/mm/yyyy…)
      - acta_text: según ACTA_MODE ("ACTA No. 243" o "243")
      - location_code: después de "DIPOL- GRISE - XXXXX" (tolera espacios/guiones)
      - recipient_cc y recipient_name desde filas inferiores (firma)
    Soporta el caso: Fila 14 "ACTA No. 243 PISO 3  DIPOL- GRISE - AMAZONAS"
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    # Blob superior (hasta ~80 filas) para patrones
    lines = []
    for r in range(1, 80):
        vals = []
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(str(v))
        if vals:
            lines.append(" | ".join(vals))
    blob = "\n".join(lines)

    # Fecha
    found_date = None
    for r in range(1, 41):
        for c in range(1, 15):
            v = ws.cell(r, c).value
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                found_date = datetime(v.year, v.month, v.day)
                break
        if found_date:
            break
    if not found_date:
        m = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", blob)
        if m:
            raw = m.group(1)
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
                try:
                    found_date = datetime.strptime(raw, fmt)
                    break
                except Exception:
                    pass
    date_str = found_date.strftime("%Y-%m-%d") if found_date else None

    # ACTA No.
    acta_no = None
    m = re.search(r"ACTA\s*No\.?\s*([A-Za-z0-9\-_/]+)", blob, flags=re.IGNORECASE)
    if m:
        acta_no = m.group(1).strip()

    if ACTA_MODE == "number_only" and acta_no:
        acta_text = acta_no
    elif acta_no:
        acta_text = f"ACTA No. {acta_no}"
    else:
        acta_text = "ACTA"

    # Ubicación: DIPOL - GRISE - XXXXX
    loc_code = None
    m = re.search(r"DIPOL\s*-\s*GRISE\s*-\s*([A-Za-zÁÉÍÓÚÑÜ0-9\s]+)", blob, flags=re.IGNORECASE)
    if m:
        loc_code = m.group(1).strip()
    # Si no se encontró en blob, intenta explícitamente en fila 14 (o 15) como reportaste
    if not loc_code:
        for r in (14, 15):
            row_text = " ".join([str(ws.cell(r, c).value) for c in range(1, 15) if ws.cell(r, c).value is not None])
            m2 = re.search(r"DIPOL\s*-\s*GRISE\s*-\s*([A-Za-zÁÉÍÓÚÑÜ0-9\s]+)", row_text, flags=re.IGNORECASE)
            if m2:
                loc_code = m2.group(1).strip()
                break

    if loc_code:
        loc_code = re.sub(r"[^A-Za-zÁÉÍÓÚÑÜ0-9\s\-]", " ", loc_code).strip()
        loc_code = re.sub(r"\s+", " ", loc_code)
        if LOCATION_MODE == "first_token":
            loc_code = loc_code.split()[0] if loc_code.split() else loc_code

    # CC y nombre (firma)
    last_rows = []
    max_row = ws.max_row
    for r in range(max_row - 40 if max_row > 40 else 1, max_row + 1):
        row_vals = []
        for c in range(1, 15):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(str(v))
        if row_vals:
            last_rows.append((r, row_vals))

    def is_cc_token(tok):
        tok = tok.replace(".", "").replace(",", "").strip()
        return tok.isdigit() and 6 <= len(tok) <= 12

    recipient_cc, recipient_name = None, None
    for r, vals in last_rows:
        tokens = []
        for cell in vals:
            tokens.extend(re.split(r"\s+", cell))
        cc_tokens = [t for t in tokens if is_cc_token(t)]
        if cc_tokens:
            recipient_cc = cc_tokens[0]
            words = [w for w in re.split(r"\s+", " ".join(vals)) if not is_cc_token(w)]
            filtered = [w for w in words if not re.search(r"CC|CEDULA|CÉDULA|DOC|IDENTIDAD", w, re.IGNORECASE)]
            line_text = " ".join(filtered).strip()
            line_text = re.sub(r"[^A-Za-zÁÉÍÓÚÑáéíóúüÜ\s\.\-]", " ", line_text)
            recipient_name = re.sub(r"\s+", " ", line_text).strip() if line_text else None
            break

    return {
        "date_str": date_str,
        "acta_text": acta_text,
        "location_code": loc_code,
        "recipient_cc": recipient_cc,
        "recipient_name": recipient_name,
    }

def read_acta_items(path, start_row=26):
    df = pd.read_excel(path, sheet_name=0, header=start_row - 1, dtype=str)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    df = df.dropna(how="all")
    return df

def build_cc_map_from_inventory(inv_xlsx):
    xl = pd.ExcelFile(inv_xlsx)
    target_sheet = None
    for name in xl.sheet_names:
        if name.strip().lower() in ["hoja cc", "cc"] or re.search(r"\bcc\b", name, re.IGNORECASE):
            target_sheet = name
            break
    if not target_sheet:
        for name in xl.sheet_names:
            if re.search(r"cc", name, re.IGNORECASE):
                target_sheet = name
                break
    if not target_sheet:
        return {}

    df = xl.parse(target_sheet, dtype=str)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip().upper() for c in df.columns]
    col_grado = next((c for c in df.columns if "GRADO" in c), None)
    col_nombre = next((c for c in df.columns if "NOMBRES" in c or ("NOMBRE" in c and "APELL" in c)), None)
    col_cc = next((c for c in df.columns if re.search(r"\bCC\b", c)), None)

    cc_map = {}
    if col_cc:
        for _, row in df.iterrows():
            cc = norm_str(row.get(col_cc))
            if not cc:
                continue
            name = norm_str(row.get(col_nombre)) if col_nombre else ""
            grado = norm_str(row.get(col_grado)) if col_grado else ""
            display = f"{grado}. {name}".strip().strip(". ")
            cc_digits = re.sub(r"\D", "", cc)
            if cc_digits:
                cc_map[cc_digits] = display if display else name or cc_digits
    return cc_map

def find_col(df, patterns):
    for col in df.columns:
        clean = re.sub(r"\s+", " ", str(col)).strip().upper()
        for pat in patterns:
            if re.search(pat, clean):
                return col
    return None

def process(inv_path, acta_path, start_row=START_ROW):
    # Carga inventario
    inv_xl = pd.ExcelFile(inv_path)
    inv_sheets = {name: inv_xl.parse(name, dtype=str) for name in inv_xl.sheet_names}

    # Metadatos acta
    meta = improved_find_acta_meta_xlsx(acta_path)
    items_df = read_acta_items(acta_path, start_row=start_row)
    cc_map = build_cc_map_from_inventory(inv_path)

    # Responsable = CC → "GRADO. NOMBRE APELLIDO"; si no, usa nombre detectado o "CC <n>"
    recipient_cc = meta["recipient_cc"]
    recipient_name = meta["recipient_name"]
    responsable_display = None
    if recipient_cc:
        responsable_display = cc_map.get(re.sub(r"\D", "", recipient_cc), None)
    if not responsable_display:
        responsable_display = (recipient_name or "").strip() or (f"CC {recipient_cc}" if recipient_cc else "SIN RESPONSABLE")

    # Preparar columnas de la tabla del acta
    col_desc = find_col(items_df, [r"DESCRIPCI[ÓO]N DEL ACTIVO", r"DESCRIPCI[ÓO]N DEL ACTIVO [ÓO] BIEN", r"DESCRIPCI[ÓO]N DEL BIEN"])
    col_desc2 = find_col(items_df, [r"DESCRIPCI[ÓO]N ADICIONAL", r"ACCESORIOS"])
    col_serie = find_col(items_df, [r"N[ÚU]MERO DE SERIE", r"N[ÚU]MERO DE SERIE DEL BIEN", r"SERIE DEL BIEN"])
    col_inv   = find_col(items_df, [r"N[ÚU]MERO INVENTARIO", r"C[ÓO]DIGO SAP", r"R6 SILOG"])
    col_valor = find_col(items_df, [r"VALOR DE ADQUISICI[ÓO]N"])
    col_cant  = find_col(items_df, [r"CANTIDAD"])

    items_work = items_df[[col_desc, col_desc2, col_serie, col_inv, col_valor, col_cant]].copy()
    items_work.columns = ["DESC", "DESC2", "SERIE", "INV", "VALOR", "CANTIDAD"]
    items_work = items_work.dropna(how="all")
    items_work["SERIE_N"] = items_work["SERIE"].map(norm_serial)

    # Mapas de serie por hoja
    def std_cols(cols): return [re.sub(r"\s+", " ", str(c)).strip().upper() for c in cols]
    sheet_cols_std = {name: std_cols(df.columns) for name, df in inv_sheets.items()}

    def col_idx(name, target):
        for i, col in enumerate(name):
            if re.search(target, col, re.IGNORECASE):
                return i
        return None

    def get_update_schema(sheet_name, cols_std):
        up = sheet_name.upper()
        if "TECNOL" in up:
            return {
                "SERIE": col_idx(cols_std, r"NUMERO DE SERIE"),
                "RESP":  col_idx(cols_std, r"\bRESPONSABLE\b"),
                "UBIC":  col_idx(cols_std, r"UBICACI[ÓO]N"),
                "ACTA":  col_idx(cols_std, r"NO\.? ACTA"),
                "FECHA": col_idx(cols_std, r"FECHA ULTIMA ASIGNACION"),
            }
        if "INMOB" in up:
            return {
                "SERIE": col_idx(cols_std, r"NUMERO DE SERIE"),
                "RESP":  col_idx(cols_std, r"\bRESPONSABLE\b"),
                "UBIC":  col_idx(cols_std, r"UBICACI[ÓO]N"),
                "ACTA":  col_idx(cols_std, r"NO\.? ACTA"),
                "FECHA": col_idx(cols_std, r"FECHA ULTIMA ASIGNACION"),
            }
        if "FUERA" in up:
            return {
                "SERIE": col_idx(cols_std, r"NUMERO DE SERIE ELEMENTO"),
                "RESP":  col_idx(cols_std, r"\bRESPONSABLE\b"),
                "UBIC":  col_idx(cols_std, r"UBICACI[ÓO]N"),
                "ACTA":  col_idx(cols_std, r"NUMERO DE ACTA|NO\.? ACTA"),
                "FECHA": col_idx(cols_std, r"FECHA ULTIMA ASIGNACION"),
            }
        return None

    schemas = {name: get_update_schema(name, sheet_cols_std[name]) for name in inv_sheets.keys()}

    updated_hits = 0
    missing_serial_or_not_found = []

    # Indexación por serie
    sheet_serial_maps = {}
    for name, df in inv_sheets.items():
        schema = schemas.get(name)
        if not schema or schema["SERIE"] is None:
            continue
        ser_col_name = df.columns[schema["SERIE"]]
        ser_map = {}
        for idx, v in df[ser_col_name].items():
            key = norm_serial(v)
            if key:
                ser_map.setdefault(key, []).append(idx)
        sheet_serial_maps[name] = ser_map

    # Aplicar actualizaciones
    for _, row in items_work.iterrows():
        serie_key = row["SERIE_N"]
        if not serie_key:
            missing_serial_or_not_found.append(("NO_SERIE", row))
            continue

        found_in_any = False
        for name, df in inv_sheets.items():
            if name not in sheet_serial_maps:
                continue
            idxs = sheet_serial_maps[name].get(serie_key, [])
            if not idxs:
                continue

            schema = schemas[name]
            for idx in idxs:
                if schema["RESP"] is not None:
                    col = df.columns[schema["RESP"]]
                    df.at[idx, col] = responsable_display
                if schema["UBIC"] is not None and meta["location_code"]:
                    col = df.columns[schema["UBIC"]]
                    df.at[idx, col] = meta["location_code"]
                if schema["ACTA"] is not None:
                    col = df.columns[schema["ACTA"]]
                    df.at[idx, col] = meta["acta_text"]
                if schema["FECHA"] is not None and meta["date_str"]:
                    col = df.columns[schema["FECHA"]]
                    df.at[idx, col] = meta["date_str"]
            updated_hits += 1
            found_in_any = True
            break

        if not found_in_any:
            missing_serial_or_not_found.append(("NOT_FOUND", row))

    # Agregar a SIN SERIAL sin borrar lo existente
    sin_serial_name = next((n for n in inv_sheets.keys() if re.search(r"SIN\s*SERIAL", n, re.IGNORECASE)), None)
    if sin_serial_name:
        ss_df = inv_sheets[sin_serial_name]
        req_cols = [
            'No',
            'DESCRIPCIÓN DEL ACTIVO Ó BIEN',
            'DESCRIPCIÓN ADICIONAL - ACCESORIOS',
            'NÚMERO DE SERIE DEL BIEN / O LOTE PARA EL CASO DE MUNICIÓN',
            'NÚMERO INVENTARIO (CÓDIGO SAP/R6 SILOG)',
            'VALOR DE ADQUISICIÓN',
            'CANTIDAD',
            'OBSERVACIONES UNIDAD',
            'OBSERVACION INTERNA',
            'No ACTA',
            'FECHA',
            'RESPONSABLE'
        ]
        for col in req_cols:
            if col not in ss_df.columns:
                ss_df[col] = pd.Series([None] * len(ss_df))

        def parse_no(x):
            try:
                return int(str(x).strip())
            except Exception:
                return None

        next_no = 1
        if len(ss_df):
            nums = [parse_no(v) for v in ss_df['No'].tolist()]
            if any(n is not None for n in nums):
                next_no = max(n for n in nums if n is not None) + 1

        append_rows = []
        for kind, r in missing_serial_or_not_found:
            desc = norm_str(r["DESC"])
            desc2 = norm_str(r["DESC2"])
            serie = norm_str(r["SERIE"])
            invn  = norm_str(r["INV"])
            valor = norm_str(r["VALOR"])
            cant  = norm_str(r["CANTIDAD"])
            append_rows.append({
                'No': next_no,
                'DESCRIPCIÓN DEL ACTIVO Ó BIEN': desc,
                'DESCRIPCIÓN ADICIONAL - ACCESORIOS': desc2,
                'NÚMERO DE SERIE DEL BIEN / O LOTE PARA EL CASO DE MUNICIÓN': serie,
                'NÚMERO INVENTARIO (CÓDIGO SAP/R6 SILOG)': invn,
                'VALOR DE ADQUISICIÓN': valor,
                'CANTIDAD': cant,
                'OBSERVACIONES UNIDAD': None,
                'OBSERVACION INTERNA': f"Auto-registro ({'SIN SERIE' if kind=='NO_SERIE' else 'SERIE NO ENCONTRADA'})",
                'No ACTA': meta["acta_text"],
                'FECHA': meta["date_str"],
                'RESPONSABLE': responsable_display,
            })
            next_no += 1

        if append_rows:
            inv_sheets[sin_serial_name] = pd.concat([ss_df, pd.DataFrame(append_rows)], ignore_index=True)

    # Guardar salida
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(inv_path))[0]
    out_path = os.path.join(os.path.dirname(inv_path), f"{base}_actualizado_{timestamp}.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in inv_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    return out_path, meta, responsable_display, updated_hits, len(missing_serial_or_not_found)

# -------- CLI / GUI --------
def main():
    parser = argparse.ArgumentParser(description="Actualizador de Inventario por Acta")
    parser.add_argument("--inventario", "-i", help="Ruta al Excel de inventario (.xlsx)")
    parser.add_argument("--acta", "-a", help="Ruta al Excel de acta (.xlsx)")
    args = parser.parse_args()

    inv_path = args.inventario
    acta_path = args.acta

    # Si no se pasan rutas por CLI, abre dialogos (en Windows/Mac suele estar Tkinter)
    if not inv_path or not acta_path:
        try:
            from tkinter import Tk, filedialog, messagebox
            root = Tk(); root.withdraw()
            if not inv_path:
                messagebox.showinfo("Actualizador", "Selecciona el archivo de INVENTARIO (.xlsx)")
                inv_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
            if not acta_path and inv_path:
                messagebox.showinfo("Actualizador", "Selecciona el archivo de ACTA (.xlsx)")
                acta_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        except Exception:
            pass

    if not inv_path or not acta_path:
        print("ERROR: Debes especificar --inventario y --acta, o seleccionarlos en el diálogo.")
        sys.exit(1)

    out_path, meta, resp, updated_count, added_count = process(inv_path, acta_path, start_row=START_ROW)
    print("\n=== RESUMEN ===")
    print(f"Archivo generado: {out_path}")
    print(f"Fecha acta: {meta.get('date_str')}")
    print(f"No. ACTA: {meta.get('acta_text')}")
    print(f"Ubicación (DIPOL-GRISE): {meta.get('location_code')}")
    print(f"Responsable asignado: {resp}")
    print(f"Actualizados por serie: {updated_count}")
    print(f"Agregados a SIN SERIAL: {added_count}")

if __name__ == "__main__":
    main()