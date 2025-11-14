# -*- coding: utf-8 -*-
"""
Actualizador de Inventario — GUI (Tkinter)
------------------------------------------
Requisitos:
  pip install pandas openpyxl
"""

import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook


# -------- Config por defecto --------
DEFAULT_START_ROW = 26                 # Fila donde empiezan los encabezados en la tabla del acta
DEFAULT_LOCATION_MODE = "raw"          # "raw" | "first_token"
DEFAULT_ACTA_MODE = "prefix"           # "prefix" | "number_only"


# -------- Utilidades --------
ES_MONTHS = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12
}
ES_ABBR = {1:"ENE",2:"FEB",3:"MAR",4:"ABR",5:"MAY",6:"JUN",7:"JUL",8:"AGO",9:"SEP",10:"OCT",11:"NOV",12:"DIC"}

def norm_str(x):
    if x is None:
        return ""
    # Evita que 'N/A' sea procesado como NaN o se pierda
    s = str(x)
    return s.strip()

def norm_serial(x):
    s = norm_str(x)
    return re.sub(r"\s+", "", s).upper()

def try_int(x):
    try:
        return int(str(x).strip())
    except Exception:
        return None

def format_stamp(dt: datetime) -> str:
    # "14NOV25 - 10_35"
    return f"{dt.day:02d}{ES_ABBR[dt.month]}{dt.year%100:02d} - {dt.hour:02d}_{dt.minute:02d}"

# --- util: primer valor no vacío a la derecha, en la misma fila
def first_right_value(ws, r, c_start, max_jump=6):
    for c in range(c_start+1, c_start+1+max_jump):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip():
            return ws.cell(r, c).value
    return None


def clean_location(raw: str, mode="raw") -> str:
    if not raw:
        return raw
    # 1) corta en palabras que no pertenecen a la ubicación
    STOP_AT = r"(OBJETIVO|ASIGNACI[ÓO]N|RESPONSABLES?|NOMBRES?|INFORMACI[ÓO]N\s+P[ÚU]BLICA|FIRMA|CARGO)"
    raw = re.split(STOP_AT, raw, flags=re.IGNORECASE)[0]

    # 2) limpia separadores y basura
    dash = r"[-–—]"
    raw = re.sub(rf"\bOBJETIVO\b.*$", "", raw, flags=re.IGNORECASE)  # redundante por seguridad
    raw = re.sub(rf"(?:\s*{dash}\s*)+$", "", raw)                     # guiones al final
    raw = re.sub(r"[|:;,]+$", "", raw)                                # otros separadores finales
    raw = re.sub(r"[^A-Za-zÁÉÍÓÚÑÜ0-9\s\-]", " ", raw)                # caracteres raros
    raw = re.sub(r"\s+", " ", raw).strip()

    # 3) si sigue muy largo, quédate con las primeras 2–4 palabras útiles
    tokens = raw.split()
    if len(tokens) >= 5:
        raw = " ".join(tokens[:4])   # ajusta 3/4 si prefieres más corto

    if mode == "first_token":
        raw = tokens[0] if tokens else raw

    return raw


# --- Fecha del acta en fila 8 (cajas: DD / MM / AA o AÑO/ANIO)
def parse_row8_date(ws):
    r = 8
    max_c = ws.max_column
    day = mon = year = None

    def cell_txt(rr, cc):
        v = ws.cell(rr, cc).value
        return str(v).strip().upper() if v is not None else ""

    for c in range(1, max_c+1):
        t = cell_txt(r, c)

        # DD
        if t in ("DD", "DIA", "DÍA"):
            raw = first_right_value(ws, r, c)
            try:
                d = int(str(raw).strip())
                if 1 <= d <= 31:
                    day = d
            except Exception:
                pass

        # MM
        if t in ("MM", "MES"):
            raw = first_right_value(ws, r, c)
            try:
                m = int(str(raw).strip())
                if 1 <= m <= 12:
                    mon = m
            except Exception:
                pass

        # AA / AÑO / ANIO
        if t in ("AA", "AÑO", "ANIO", "AÑO"):
            raw = first_right_value(ws, r, c)
            if raw is not None:
                s = str(raw).strip()
                if s.isdigit():
                    y = int(s)
                    if 0 <= y <= 99:
                        year = 2000 + y
                    elif 1900 <= y <= 2100:
                        year = y

    # Fallback: todos los números de la fila 8, en orden de aparición
    if day is None or mon is None or year is None:
        nums = []
        for c in range(1, max_c+1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s.isdigit():
                nums.append((int(s), c))
        for i in range(len(nums)):
            for j in range(i+1, len(nums)):
                for k in range(j+1, len(nums)):
                    d, _ = nums[i]
                    m, _ = nums[j]
                    y, _ = nums[k]
                    if 1 <= d <= 31 and 1 <= m <= 12:
                        if 0 <= y <= 99:
                            y = 2000 + y
                        if 1900 <= y <= 2100:
                            day, mon, year = d, m, y
                            break
                if day and mon and year:
                    break
            if day and mon and year:
                break

    return datetime(year, mon, day) if (day and mon and year) else None


# --- Responsable: detectar en tabla de ASISTENTES (CÉDULA / NOMBRES / CARGO)
# --- Responsable: detectar en tabla de ASISTENTES (GRADO / CÉDULA / NOMBRES / CARGO)
def find_responsable(ws):
    max_row, max_col = ws.max_row, ws.max_column

    def norm_cell(v):
        return str(v).strip().upper() if v is not None else ""

    # 1) Localizar fila de encabezados
    header_row = None
    col_idx = {"GRADO": None, "CEDULA": None, "NOMBRES": None, "CARGO": None}

    for r in range(1, min(max_row, 400) + 1):
        row_labels = [norm_cell(ws.cell(r, c).value) for c in range(1, min(max_col, 60) + 1)]
        row_text = " | ".join(row_labels)

        has_ced = ("CÉDULA" in row_text) or ("CEDULA" in row_text)
        has_nom = ("NOMBRES" in row_text) or ("NOMBRES Y APELLIDOS" in row_text)
        has_car = ("CARGO" in row_text)

        if has_ced and has_nom and has_car:
            header_row = r
            for c, lab in enumerate(row_labels, start=1):
                if re.search(r"\bGRADO\b", lab):
                    col_idx["GRADO"] = c
                if re.search(r"\bC[ÉE]DULA\b", lab):
                    col_idx["CEDULA"] = c
                if re.search(r"\bNOMBRES(\s+Y\s+APELLIDOS)?\b", lab):
                    col_idx["NOMBRES"] = c
                if re.search(r"\bCARGO\b", lab):
                    col_idx["CARGO"] = c
            break

    # 2) Buscar fila con CARGO = "FUNCIONARIO QUE RECIBE"
    if header_row and col_idx["CEDULA"] and col_idx["NOMBRES"] and col_idx["CARGO"]:
        for r in range(header_row + 1, min(header_row + 120, max_row) + 1):
            cargo = norm_cell(ws.cell(r, col_idx["CARGO"]).value)
            if re.search(r"\bFUNCIONARIO\s+QUE\s+RECIBE\b", cargo, re.IGNORECASE):
                # extraer CC, NOMBRE y (opcional) GRADO
                cc_raw = ws.cell(r, col_idx["CEDULA"]).value
                name_raw = ws.cell(r, col_idx["NOMBRES"]).value
                grade_raw = ws.cell(r, col_idx["GRADO"]).value if col_idx["GRADO"] else None

                cc = None
                if cc_raw is not None:
                    d = re.sub(r"\D", "", str(cc_raw))
                    if d.isdigit() and 6 <= len(d) <= 12:
                        cc = d

                name = None
                if name_raw is not None:
                    s = str(name_raw).strip()
                    name = re.sub(r"\s+", " ", s) or None

                grade = None
                if grade_raw is not None:
                    g = str(grade_raw).strip()
                    grade = re.sub(r"\s+", " ", g).upper() or None

                return cc, name, grade

    # 3) Respaldo: ventana alrededor del rótulo (sin grado garantizado)
    patt = re.compile(r"FUNCIONARIO\s+QUE\s+RECIBE", re.IGNORECASE)
    label_r = label_c = None
    for r in range(1, min(max_row, 200) + 1):
        for c in range(1, min(max_col, 50) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and patt.search(v):
                label_r, label_c = r, c
                break
        if label_r:
            break

    if not label_r:
        return None, None, None

    def harvest_window(rr, cc_label, left_cols=12, right_cols=12, rows_up=1, rows_down=2):
        r0 = max(1, rr - rows_up)
        r1 = min(max_row, rr + rows_down)
        cL = max(1, cc_label - left_cols)
        cR = min(max_col, cc_label + right_cols)

        left_texts, right_texts = [], []
        left_digits, right_digits = [], []

        for r in range(r0, r1 + 1):
            for c in range(cL, cc_label):  # izquierda
                s = norm_cell(ws.cell(r, c).value)
                if not s:
                    continue
                left_texts.append(s)
                d = re.sub(r"\D", "", s)
                if d.isdigit() and 6 <= len(d) <= 12:
                    left_digits.append(d)
            for c in range(cc_label + 1, cR + 1):  # derecha
                s = norm_cell(ws.cell(r, c).value)
                if not s:
                    continue
                right_texts.append(s)
                d = re.sub(r"\D", "", s)
                if d.isdigit() and 6 <= len(d) <= 12:
                    right_digits.append(d)

        return (left_texts, left_digits, right_texts, right_digits)

    left_texts, left_digits, right_texts, right_digits = harvest_window(label_r, label_c)
    cc = left_digits[0] if left_digits else (right_digits[0] if right_digits else None)

    def clean_name(txts):
        joined = " ".join(txts)
        joined = re.sub(r"(CC|C[ÉE]DULA|DOC(?:UMENTO)?|IDENTIDAD|N[°O]\.?)\s*[:\-]?", " ", joined, flags=re.IGNORECASE)
        joined = re.sub(r"\d{6,}", " ", joined)
        joined = re.sub(r"[^A-Za-zÁÉÍÓÚÑáéíóúüÜ\s\.\-]", " ", joined)
        return re.sub(r"\s+", " ", joined).strip() or None

    nombre = clean_name(right_texts) or clean_name(left_texts)
    grade = None  # en el respaldo no es fiable detectar grado
    return cc, nombre, grade



# --- Localiza la fila donde aparece el final del listado ("OBSERVACIONES Y RECOMENDACIONES")
def find_end_marker_row(acta_path):
    wb = load_workbook(acta_path, data_only=True)
    ws = wb.worksheets[0]
    max_row = ws.max_row
    patt = re.compile(r"OBSERVACIONES\s+Y\s+RECOMENDACIONES", re.IGNORECASE)
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and patt.search(v):
                return r
    return None


def improved_find_acta_meta_xlsx(path, location_mode=DEFAULT_LOCATION_MODE, acta_mode=DEFAULT_ACTA_MODE):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    # === Fecha exacta desde fila 8 (DD/MM/AA)
    found_date = parse_row8_date(ws)

    # Construyo blob para patrones (ACTA, ubicación)
    lines = []
    for r in range(1, 80):
        vals = []
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(str(v))
        if vals:
            lines.append(" | ".join(vals))
    blob = "\n".join(lines)

    date_str = found_date.strftime("%Y-%m-%d") if found_date else None

    # === ACTA No.
    acta_no = None
    m = re.search(r"ACTA\s*No\.?\s*([A-Za-z0-9\-_/]+)", blob, flags=re.IGNORECASE)
    if m:
        acta_no = m.group(1).strip()
    if acta_mode == "number_only" and acta_no:
        acta_text = acta_no
    elif acta_no:
        acta_text = f"ACTA No. {acta_no}"
    else:
        acta_text = "ACTA"

    # === Ubicación DIPOL - GRISE - XXX (cortar estrictamente en "OBJETIVO")
    loc_code = None
    dash = r"[-–—]"
    loc_pat = rf"DIPOL\s*{dash}\s*GRISE\s*{dash}\s*([A-Za-zÁÉÍÓÚÑÜ0-9\s\-\:\|\.\,]+)"

    m = re.search(loc_pat, blob, flags=re.IGNORECASE)
    if m:
        loc_code = m.group(1).strip()
    if not loc_code:
        for r in (14, 15):
            row_text = " ".join([str(ws.cell(r, c).value) for c in range(1, 15) if ws.cell(r, c).value is not None])
            m2 = re.search(loc_pat, row_text, flags=re.IGNORECASE)
            if m2:
                loc_code = m2.group(1).strip()
                break

    if loc_code:
        # 1) Dejar estrictamente lo anterior a "OBJETIVO"
        loc_code = re.split(r"\bOBJETIVO\b", loc_code, flags=re.IGNORECASE)[0]

        # 2) Limpieza de separadores finales y espacios
        loc_code = re.sub(r"[|:;,]+$", "", loc_code)
        loc_code = re.sub(rf"(?:\s*{dash}\s*)+$", "", loc_code)  # guiones al final
        loc_code = re.sub(r"[^A-Za-zÁÉÍÓÚÑÜ0-9\s\-]", " ", loc_code)
        loc_code = re.sub(r"\s+", " ", loc_code).strip()

        # 3) first_token si se solicita
        if location_mode == "first_token":
            parts = loc_code.split()
            loc_code = parts[0] if parts else loc_code

    # === Responsable (solo CC; el nombre final se resuelve con Hoja CC en process_inventory)
    recipient_cc, recipient_name, recipient_grade = find_responsable(ws)

    return {
        "date_str": date_str,
        "acta_text": acta_text,
        "location_code": loc_code,
        "recipient_cc": recipient_cc,
        "recipient_name": recipient_name,
        "recipient_grade": recipient_grade,
    }





def build_cc_map_from_inventory(inv_xlsx):
    xl = pd.ExcelFile(inv_xlsx)
    target_sheet = None
    for name in xl.sheet_names:
        if name.strip().lower() in ["hoja cc", "cc"] or re.search(r"\bcc\b", name, re.IGNORECASE):
            target_sheet = name
            break
    if not target_sheet:
        for name in xl.sheet_names:
            if re.search(r"cc", name, re.IGNORECASE):
                target_sheet = name
                break
    if not target_sheet:
        return {}

    # Mantener 'N/A' literal
    df = xl.parse(target_sheet, dtype=str, keep_default_na=False)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip().upper() for c in df.columns]
    col_grado = next((c for c in df.columns if "GRADO" in c), None)
    col_nombre = next((c for c in df.columns if "NOMBRES" in c or ("NOMBRE" in c and "APELL" in c)), None)
    col_cc = next((c for c in df.columns if re.search(r"\bCC\b", c)), None)

    cc_map = {}
    if col_cc:
        for _, row in df.iterrows():
            cc = (row.get(col_cc) if col_cc else "") or ""
            name = (row.get(col_nombre) if col_nombre else "") or ""
            grado = (row.get(col_grado) if col_grado else "") or ""
            display = f"{str(grado).strip()}. {str(name).strip()}".strip().strip(". ")
            cc_digits = re.sub(r"\D", "", str(cc))
            if cc_digits:
                cc_map[cc_digits] = display if display else (str(name).strip() or cc_digits)
    return cc_map


# --- Lee la tabla de la ACTA y corta en el marcador de fin
def read_acta_items(path, start_row=DEFAULT_START_ROW):
    # Localiza la fila del marcador
    end_marker_row = find_end_marker_row(path)
    # nrows = cantidad de filas de datos por debajo del encabezado, hasta (antes de) el marcador
    nrows = None
    if end_marker_row and end_marker_row > start_row:
        nrows = (end_marker_row - 1) - start_row  # antes del rótulo

    # Mantener 'N/A' literal
    df = pd.read_excel(path, sheet_name=0, header=start_row - 1, dtype=str, nrows=nrows, keep_default_na=False)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    # no dropna(how="all") para no perder filas con "N/A"; pero sí eliminar filas realmente vacías
    df = df[~df.isna().all(axis=1)]
    return df


def find_col(df, patterns):
    for col in df.columns:
        clean = re.sub(r"\s+", " ", str(col)).strip().upper()
        for pat in patterns:
            if re.search(pat, clean):
                return col
    return None


def process_inventory(inv_path, acta_path, start_row, location_mode, acta_mode, log):
    log("Cargando inventario...\n")
    inv_xl = pd.ExcelFile(inv_path)

    # Leer TODAS las hojas manteniendo 'N/A'
    inv_sheets = {name: inv_xl.parse(name, dtype=str, keep_default_na=False) for name in inv_xl.sheet_names}

    log("Leyendo metadatos del acta...\n")
    meta = improved_find_acta_meta_xlsx(acta_path, location_mode, acta_mode)

    log(f"Fecha: {meta.get('date_str')}\n")
    log(f"ACTA: {meta.get('acta_text')}\n")
    log(f"Ubicación: {meta.get('location_code')}\n")
    log(f"FUNCIONARIO QUE RECIBE — CC: {meta.get('recipient_cc')} | Nombre: {meta.get('recipient_name')}\n")

    log("Construyendo mapa CC -> 'GRADO. NOMBRE APELLIDO'...\n")
    cc_map = build_cc_map_from_inventory(inv_path)

    # --- Responsable display: usa GRADO + NOMBRE si vienen del acta; si no, resuelve por CC en Hoja CC
    cc_raw = meta.get("recipient_cc")
    responsable_display = ""

    if cc_raw:
            cc_digits = re.sub(r"\D", "", str(cc_raw))
            if cc_digits:
                # cc_map ya viene como "GRADO. NOMBRE"
                responsable_display = cc_map.get(cc_digits, "")

        # Fallbacks: si no hubo match en Hoja CC
    if not responsable_display:
            # último recurso legible
            responsable_display = "SIN RESPONSABLE"

    log("Leyendo ítems del acta...\n")
    items_df = read_acta_items(acta_path, start_row=start_row)

    # --- Columnas del acta, incluyendo OBSERVACIONES
    col_desc = find_col(items_df, [r"DESCRIPCI[ÓO]N DEL ACTIVO", r"DESCRIPCI[ÓO]N DEL ACTIVO [ÓO] BIEN", r"DESCRIPCI[ÓO]N DEL BIEN"])
    col_desc2 = find_col(items_df, [r"DESCRIPCI[ÓO]N ADICIONAL", r"ACCESORIOS"])
    col_serie = find_col(items_df, [r"N[ÚU]MERO DE SERIE", r"N[ÚU]MERO DE SERIE DEL BIEN", r"SERIE DEL BIEN"])
    col_inv   = find_col(items_df, [r"N[ÚU]MERO INVENTARIO", r"C[ÓO]DIGO SAP", r"R6 SILOG"])
    col_valor = find_col(items_df, [r"VALOR DE ADQUISICI[ÓO]N"])
    col_cant  = find_col(items_df, [r"CANTIDAD"])
    col_obs   = find_col(items_df, [r"\bOBSERVACION(?:ES)?\b", r"\bOBSERVACIONES DEL ELEMENTO\b", r"\bOBSERVACIONES\b"])

    # Conservar literales incluyendo "N/A"
    use_cols = [col_desc, col_desc2, col_serie, col_inv, col_valor, col_cant, col_obs]
    items_work = items_df[use_cols].copy()
    items_work.columns = ["DESC", "DESC2", "SERIE", "INV", "VALOR", "CANTIDAD", "OBS"]
    items_work["SERIE_N"] = items_work["SERIE"].map(norm_serial)

    def std_cols(cols): return [re.sub(r"\s+", " ", str(c)).strip().upper() for c in cols]
    sheet_cols_std = {name: std_cols(df.columns) for name, df in inv_sheets.items()}

    def col_idx(name, target):
        for i, col in enumerate(name):
            if re.search(target, col, re.IGNORECASE):
                return i
        return None

    # --- Esquemas por hoja: añadimos OBSERVACIONES UNIDAD
    def get_update_schema(sheet_name, cols_std):
        up = sheet_name.upper()
        schema_common = {
            "SERIE": col_idx(cols_std, r"NUMERO DE SERIE"),
            "RESP":  col_idx(cols_std, r"\bRESPONSABLE\b"),
            "UBIC":  col_idx(cols_std, r"UBICACI[ÓO]N"),
            "ACTA":  col_idx(cols_std, r"(NO\.?\s*ACTA|NUMERO DE ACTA)"),
            "FECHA": col_idx(cols_std, r"FECHA ULTIMA ASIGNACION"),
            "OBS_UNIT": col_idx(cols_std, r"OBSERVACIONES? UNIDAD")
        }
        if "FUERA" in up:
            schema_common["SERIE"] = schema_common["SERIE"] or col_idx(cols_std, r"NUMERO DE SERIE ELEMENTO")
            schema_common["ACTA"]  = schema_common["ACTA"] or col_idx(cols_std, r"NUMERO DE ACTA|NO\.?\s*ACTA")
        return schema_common

    schemas = {name: get_update_schema(name, sheet_cols_std[name]) for name in inv_sheets.keys()}

    log("Indexando inventario por número de serie...\n")
    sheet_serial_maps = {}
    for name, df in inv_sheets.items():
        schema = schemas.get(name)
        if not schema or schema["SERIE"] is None:
            continue
        ser_col_name = df.columns[schema["SERIE"]]
        ser_map = {}
        for idx, v in df[ser_col_name].items():
            key = norm_serial(v)
            if key:
                ser_map.setdefault(key, []).append(idx)
        sheet_serial_maps[name] = ser_map

    updated_hits = 0
    missing_serial_or_not_found = []

    log("Aplicando actualizaciones...\n")
    for _, row in items_work.iterrows():
        serie_key = row["SERIE_N"]
        obs_text  = norm_str(row["OBS"]) or None  # conservar "N/A" como texto
        if not serie_key:
            missing_serial_or_not_found.append(("NO_SERIE", row))
            continue

        found_in_any = False
        for name, df in inv_sheets.items():
            if name not in sheet_serial_maps:
                continue
            idxs = sheet_serial_maps[name].get(serie_key, [])
            if not idxs:
                continue

            schema = schemas[name]
            for idx in idxs:
                if schema["RESP"] is not None:
                    col = df.columns[schema["RESP"]]
                    df.at[idx, col] = responsable_display
                if schema["UBIC"] is not None and meta["location_code"]:
                    col = df.columns[schema["UBIC"]]
                    df.at[idx, col] = meta["location_code"]
                if schema["ACTA"] is not None:
                    col = df.columns[schema["ACTA"]]
                    df.at[idx, col] = meta["acta_text"]
                if schema["FECHA"] is not None and meta["date_str"]:
                    col = df.columns[schema["FECHA"]]
                    df.at[idx, col] = meta["date_str"]
                if schema["OBS_UNIT"] is not None and obs_text:
                    col = df.columns[schema["OBS_UNIT"]]
                    df.at[idx, col] = obs_text
            updated_hits += 1
            found_in_any = True
            break

        if not found_in_any:
            missing_serial_or_not_found.append(("NOT_FOUND", row))

    # --- Hoja SIN SERIAL: agregar y llevar observaciones a "OBSERVACIONES UNIDAD"
    sin_serial_name = next((n for n in inv_sheets.keys() if re.search(r"SIN\s*SERIAL", n, re.IGNORECASE)), None)
    if sin_serial_name:
        ss_df = inv_sheets[sin_serial_name]
        req_cols = [
            'No',
            'DESCRIPCIÓN DEL ACTIVO Ó BIEN',
            'DESCRIPCIÓN ADICIONAL - ACCESORIOS',
            'NÚMERO DE SERIE DEL BIEN / O LOTE PARA EL CASO DE MUNICIÓN',
            'NÚMERO INVENTARIO (CÓDIGO SAP/R6 SILOG)',
            'VALOR DE ADQUISICIÓN',
            'CANTIDAD',
            'OBSERVACIONES UNIDAD',
            'OBSERVACION INTERNA',
            'No ACTA',
            'FECHA',
            'RESPONSABLE'
        ]
        for col in req_cols:
            if col not in ss_df.columns:
                ss_df[col] = pd.Series([None] * len(ss_df))

        def parse_no(x):
            try:
                return int(str(x).strip())
            except Exception:
                return None

        next_no = 1
        if len(ss_df):
            nums = [parse_no(v) for v in ss_df['No'].tolist()]
            if any(n is not None for n in nums):
                next_no = max(n for n in nums if n is not None) + 1

        append_rows = []
        for kind, r in missing_serial_or_not_found:
            desc = norm_str(r["DESC"])      # conserva "N/A"
            desc2 = norm_str(r["DESC2"])    # conserva "N/A"
            serie = norm_str(r["SERIE"])    # conserva "N/A"
            invn  = norm_str(r["INV"])      # conserva "N/A"
            valor = norm_str(r["VALOR"])    # conserva "N/A"
            cant  = norm_str(r["CANTIDAD"])
            obs   = norm_str(r["OBS"]) or None
            append_rows.append({
                'No': next_no,
                'DESCRIPCIÓN DEL ACTIVO Ó BIEN': desc,
                'DESCRIPCIÓN ADICIONAL - ACCESORIOS': desc2,
                'NÚMERO DE SERIE DEL BIEN / O LOTE PARA EL CASO DE MUNICIÓN': serie,
                'NÚMERO INVENTARIO (CÓDIGO SAP/R6 SILOG)': invn,
                'VALOR DE ADQUISICIÓN': valor,
                'CANTIDAD': cant,
                'OBSERVACIONES UNIDAD': obs,  # <<<<<< observaciones del acta (incluye "N/A" literal)
                'OBSERVACION INTERNA': f"Auto-registro ({'SIN SERIE' if kind=='NO_SERIE' else 'SERIE NO ENCONTRADA'})",
                'No ACTA': meta["acta_text"],
                'FECHA': meta["date_str"],
                'RESPONSABLE': responsable_display,
            })
            next_no += 1

        if append_rows:
            inv_sheets[sin_serial_name] = pd.concat([ss_df, pd.DataFrame(append_rows)], ignore_index=True)

    # --- Guardar con el formato "14NOV25 - 10_35"
    stamp = format_stamp(datetime.now())
    base = os.path.splitext(os.path.basename(inv_path))[0]
    out_name = f"{base} {stamp}.xlsx"
    out_path = os.path.join(os.path.dirname(inv_path), out_name)

    log(f"Guardando archivo: {out_path}\n")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in inv_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    return out_path, meta, responsable_display, updated_hits, len(missing_serial_or_not_found)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Actualizador de Inventario — GRISE")
        self.geometry("900x600")
        self.minsize(880, 560)

        self.inv_path = tk.StringVar()
        self.acta_path = tk.StringVar()

        self.start_row = tk.IntVar(value=DEFAULT_START_ROW)
        self.location_mode = tk.StringVar(value=DEFAULT_LOCATION_MODE)
        self.acta_mode = tk.StringVar(value=DEFAULT_ACTA_MODE)

        self.meta_fecha = tk.StringVar(value="-")
        self.meta_acta  = tk.StringVar(value="-")
        self.meta_ubic  = tk.StringVar(value="-")
        self.meta_cc    = tk.StringVar(value="-")
        self.meta_name  = tk.StringVar(value="-")

        self._build_ui()

    def _build_ui(self):
        pad = {'padx': 10, 'pady': 6}

        frm_files = ttk.LabelFrame(self, text="Archivos")
        frm_files.pack(fill="x", **pad)

        ttk.Label(frm_files, text="Inventario (.xlsx):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(frm_files, textvariable=self.inv_path).grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(frm_files, text="Buscar...", command=self.pick_inventory).grid(row=0, column=2, padx=8, pady=6)

        ttk.Label(frm_files, text="Acta (.xlsx):").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(frm_files, textvariable=self.acta_path).grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(frm_files, text="Buscar...", command=self.pick_acta).grid(row=1, column=2, padx=8, pady=6)

        frm_files.columnconfigure(1, weight=1)

        frm_opts = ttk.LabelFrame(self, text="Opciones de procesamiento")
        frm_opts.pack(fill="x", **pad)

        ttk.Label(frm_opts, text="Fila inicio tabla (ACTA):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Spinbox(frm_opts, from_=1, to=200, textvariable=self.start_row, width=8).grid(row=0, column=1, sticky="w", padx=8, pady=6)

        ttk.Label(frm_opts, text="Formato No. ACTA:").grid(row=0, column=2, sticky="w", padx=8, pady=6)
        cbo_acta = ttk.Combobox(frm_opts, textvariable=self.acta_mode, values=("prefix", "number_only"), state="readonly", width=14)
        cbo_acta.grid(row=0, column=3, sticky="w", padx=8, pady=6)
        ttk.Label(frm_opts, text="(prefix= 'ACTA No. 243', number_only='243')").grid(row=0, column=4, sticky="w")

        ttk.Label(frm_opts, text="Ubicación (DIPOL-GRISE):").grid(row=1, column=2, sticky="w", padx=8, pady=6)
        cbo_loc = ttk.Combobox(frm_opts, textvariable=self.location_mode, values=("raw", "first_token"), state="readonly", width=14)
        cbo_loc.grid(row=1, column=3, sticky="w", padx=8, pady=6)
        ttk.Label(frm_opts, text="(raw = completa, first_token = 1ra palabra)").grid(row=1, column=4, sticky="w")

        frm_meta = ttk.LabelFrame(self, text="Metadatos detectados del ACTA")
        frm_meta.pack(fill="x", **pad)

        ttk.Label(frm_meta, text="Fecha:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ttk.Label(frm_meta, textvariable=self.meta_fecha).grid(row=0, column=1, sticky="w", padx=8, pady=4)

        ttk.Label(frm_meta, text="No. ACTA:").grid(row=0, column=2, sticky="w", padx=8, pady=4)
        ttk.Label(frm_meta, textvariable=self.meta_acta).grid(row=0, column=3, sticky="w", padx=8, pady=4)

        ttk.Label(frm_meta, text="Ubicación:").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ttk.Label(frm_meta, textvariable=self.meta_ubic).grid(row=1, column=1, sticky="w", padx=8, pady=4)

        ttk.Label(frm_meta, text="CC receptor:").grid(row=1, column=2, sticky="w", padx=8, pady=4)
        ttk.Label(frm_meta, textvariable=self.meta_cc).grid(row=1, column=3, sticky="w", padx=8, pady=4)

        ttk.Label(frm_meta, text="Nombre receptor:").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        ttk.Label(frm_meta, textvariable=self.meta_name).grid(row=2, column=1, sticky="w", padx=8, pady=4)

        for i in (1,3):
            frm_meta.columnconfigure(i, weight=1)

        frm_actions = ttk.Frame(self)
        frm_actions.pack(fill="x", **pad)

        self.btn_preview = ttk.Button(frm_actions, text="Previsualizar ACTA", command=self.preview_meta)
        self.btn_preview.pack(side="left", padx=6)

        self.btn_run = ttk.Button(frm_actions, text="Procesar y generar Excel", command=self.run_process)
        self.btn_run.pack(side="right", padx=6)

        frm_log = ttk.LabelFrame(self, text="Registro")
        frm_log.pack(fill="both", expand=True, **pad)
        self.txt = tk.Text(frm_log, height=14, wrap="word")
        self.txt.pack(fill="both", expand=True, padx=8, pady=8)

    def pick_inventory(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.inv_path.set(path)

    def pick_acta(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.acta_path.set(path)

    def log(self, msg):
        self.txt.insert("end", msg)
        self.txt.see("end")
        self.update_idletasks()

    def preview_meta(self):
        acta = self.acta_path.get().strip()
        if not acta:
            messagebox.showwarning("Falta archivo", "Selecciona el archivo de ACTA (.xlsx)")
            return
        try:
            meta = improved_find_acta_meta_xlsx(
                acta,
                location_mode=self.location_mode.get(),
                acta_mode=self.acta_mode.get()
            )

            # Mostrar fecha/acta/ubicación tal cual
            self.meta_fecha.set(meta.get("date_str") or "-")
            self.meta_acta.set(meta.get("acta_text") or "-")
            self.meta_ubic.set(meta.get("location_code") or "-")

            # Resolver responsable exclusivamente por CC contra Hoja CC (si ya escogiste inventario)
            cc_show = meta.get("recipient_cc") or "-"
            self.meta_cc.set(cc_show)

            resolved_name = "-"
            inv = self.inv_path.get().strip()
            if inv and meta.get("recipient_cc"):
                try:
                    cc_map = build_cc_map_from_inventory(inv)
                    cc_digits = re.sub(r"\D", "", str(meta["recipient_cc"]))
                    if cc_digits and cc_digits in cc_map:
                        resolved_name = cc_map[cc_digits]  # "GRADO. NOMBRE"
                except Exception:
                    pass

            # Muestra el nombre resuelto (si no hay inv seleccionado o no coincide, queda "-")
            self.meta_name.set(resolved_name)

            self.log("Metadatos del ACTA actualizados (responsable resuelto por CC en Hoja CC si es posible).\n")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el ACTA.\n\n{e}")


    def run_process(self):
        inv = self.inv_path.get().strip()
        acta = self.acta_path.get().strip()
        if not inv or not acta:
            messagebox.showwarning("Faltan archivos", "Selecciona el Excel de INVENTARIO y el de ACTA.")
            return

        try:
            self.txt.delete("1.0", "end")
            self.log("Iniciando procesamiento...\n")

            out_path, meta, resp, updated_count, added_count = process_inventory(
                inv_path=inv,
                acta_path=acta,
                start_row=int(self.start_row.get()),
                location_mode=self.location_mode.get(),
                acta_mode=self.acta_mode.get(),
                log=self.log
            )

            self.log("\n=== RESUMEN ===\n")
            self.log(f"Archivo generado: {out_path}\n")
            self.log(f"Fecha acta: {meta.get('date_str')}\n")
            self.log(f"No. ACTA: {meta.get('acta_text')}\n")
            self.log(f"Ubicación: {meta.get('location_code')}\n")
            self.log(f"Responsable (FUNCIONARIO QUE RECIBE): {resp}\n")
            self.log(f"Actualizados por serie: {updated_count}\n")
            self.log(f"Agregados a SIN SERIAL: {added_count}\n")

            if messagebox.askyesno("Listo", f"Archivo generado:\n{out_path}\n\n¿Abrir la carpeta contenedora?"):
                os.startfile(os.path.dirname(out_path))

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error durante el proceso.\n\n{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
